### checkMask() method credits 'https://github.com/balajisrinivas/Face-Mask-Detection' ### 


import msvcrt
from tensorflow.keras.applications.mobilenet_v2 import preprocess_input
from tensorflow.keras.preprocessing.image import img_to_array
from tensorflow.keras.models import load_model
from imutils.video import VideoStream
import numpy as np
import imutils
import time
import cv2
import os
import sys
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import os.path
from os import path
from datetime import date
import tkinter as tk
from tkinter import messagebox as mb
from PIL import Image, ImageTk
import tkinter.font as tkFont
import cv2 
import time
s = "Welcome to IIITN Attendance Portal.";


filename = "Student_database.xlsx"
sheetlist =["HCI","CN","TOC","DBMS","TC"]

detailsList2 = ["Confirming identity..."]

identityConfirmed = False


def Date(sheet1):    
	date1= "/10/2020"    
	for i in range(31):
		x=str(i+1)+ date1
		y= sheet1.cell(row=1,column =i+3)
		y.value = x

def create_excel():
	wb = Workbook()
	ws = wb.active
	if(path.exists(filename) == False):
		wb.save(filename)       

		for j in range(len(sheetlist)):
			wb = load_workbook(filename)
			sheet1 = wb.create_sheet(sheetlist[j],-1)
			sheet1["A1"] = "ID"
			sheet1["B1"] = "Name"

			Date(sheet1)
			
			sheet1["A2"] = "BT18CSE052"
			sheet1["A3"] = "BT18CSE054"
			sheet1["A4"] = "BT18CSE099"
			sheet1["A5"] = "BT18CSE101"
			sheet1["A6"] = "BT18CSE102"

			sheet1["B2"] = "Pulkit Batra"
			sheet1["B3"] = "Gopal Pandey"
			sheet1["B4"] = "Nimish Palinkar"
			sheet1["B5"] = "Shubham Sagar"
			sheet1["B6"] = "K. Bhanu Prakash Reddy"

			dims = {}
			for row in sheet1.rows:
				for cell in row:
					if cell.value:
						 dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
			for col, value in dims.items():
				sheet1.column_dimensions[col].width = value + 1
					
			sheet1.freeze_panes = "C2"
			wb.save(filename)
			
		std = wb['Sheet']
		wb.remove(std)    
		wb.save(filename)     


def Trigger(id,subject):
	if (path.exists("Student_database.xlsx") == True):
		wb = Workbook()
		ws = wb.active
		wb = load_workbook(filename)
		
		sheet1 = wb[subject]
		for i in range(sheet1.max_row):
			y=sheet1.cell(row = i+1,column=1)
			if y.value == id:                
				today = date.today()
				today = today.strftime("%d/%m/%Y")
				for j in range(sheet1.max_column):
					x = sheet1.cell(row = 1,column=j+3)
					if x.value == today:
						z =sheet1.cell(row = i+1,column=j+3)
						test="P"
						z.value = test
						wb.save(filename)
						a = sheet1.cell(row = i+1,column = 2)
						return a.value
		
	else:
		create_excel()
		wb = Workbook()
		ws = wb.active
		wb = load_workbook(filename)
		
		sheet1 = wb[subject]
		for i in range(sheet1.max_row):
			y=sheet1.cell(row = i+1,column=1)
			if y.value == id:                
				today = date.today()
				today = today.strftime("%d/%m/%Y")
				for j in range(sheet1.max_column):
					x = sheet1.cell(row = 1,column=j+3)
					if x.value == today:
						z =sheet1.cell(row = i+1,column=j+3)
						test="P"
						z.value = test
						wb.save(filename)
						a = sheet1.cell(row = i+1,column = 2)
						return a.value
					  

def detect_and_predict_mask(frame, faceNet, maskNet):
	# grab the dimensions of the frame and then construct a blob
	# from it
	(h, w) = frame.shape[:2]
	blob = cv2.dnn.blobFromImage(frame, 1.0, (224, 224),
		(104.0, 177.0, 123.0))

	# pass the blob through the network and obtain the face detections
	faceNet.setInput(blob)
	detections = faceNet.forward()
	#print(detections.shape)

	# initialize our list of faces, their corresponding locations,
	# and the list of predictions from our face mask network
	faces = []
	locs = []
	preds = []

	# loop over the detections
	for i in range(0, detections.shape[2]):
		# extract the confidence (i.e., probability) associated with
		# the detection
		confidence = detections[0, 0, i, 2]

		# filter out weak detections by ensuring the confidence is
		# greater than the minimum confidence
		if confidence > 0.5:
			# compute the (x, y)-coordinates of the bounding box for
			# the object
			box = detections[0, 0, i, 3:7] * np.array([w, h, w, h])
			(startX, startY, endX, endY) = box.astype("int")

			# ensure the bounding boxes fall within the dimensions of
			# the frame
			(startX, startY) = (max(0, startX), max(0, startY))
			(endX, endY) = (min(w - 1, endX), min(h - 1, endY))

			# extract the face ROI, convert it from BGR to RGB channel
			# ordering, resize it to 224x224, and preprocess it
			face = frame[startY:endY, startX:endX]
			face = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
			face = cv2.resize(face, (224, 224))
			face = img_to_array(face)
			face = preprocess_input(face)

			# add the face and bounding boxes to their respective
			# lists
			faces.append(face)
			locs.append((startX, startY, endX, endY))

	# only make a predictions if at least one face was detected
	if len(faces) > 0:
		# for faster inference we'll make batch predictions on *all*
		# faces at the same time rather than one-by-one predictions
		# in the above `for` loop
		faces = np.array(faces, dtype="float32")
		preds = maskNet.predict(faces, batch_size=32)

	# return a 2-tuple of the face locations and their corresponding
	# locations
	return (locs, preds)


def checkMask():

	countMask = 0
	countNotmask = 0
	flag = 0
	flagm = 0

	# load our serialized face detector model from disk
	prototxtPath = r"face_detector\deploy.prototxt"
	weightsPath = r"face_detector\res10_300x300_ssd_iter_140000.caffemodel"
	faceNet = cv2.dnn.readNet(prototxtPath, weightsPath)

	# load the face mask detector model from disk
	maskNet = load_model("mask_detector.model")

	# initialize the video stream
	x = "Checking for mask ....."
	print(x)
	vs = VideoStream(src=0).start()

	# loop over the frames from the video stream
	while True:     

		# grab the frame from the threaded video stream and resize it
		# to have a maximum width of 400 pixels
		frame = vs.read()
		frame = imutils.resize(frame, width=400)

		# detect faces in the frame and determine if they are wearing a
		# face mask or not
		(locs, preds) = detect_and_predict_mask(frame, faceNet, maskNet)

		# loop over the detected face locations and their corresponding
		# locations
		for (box, pred) in zip(locs, preds):
			# unpack the bounding box and predictions
			(startX, startY, endX, endY) = box
			(mask, withoutMask) = pred

			# determine the class label and color we'll use to draw
			# the bounding box and text
			label = "" 
			if mask > withoutMask:
				label="Mask"
				countMask = countMask + 1
				#print(countMask)
				if (countMask>10):
					flagm=1 
					flag=1              
					break
			else:
				label="No Mask"
				countNotmask = countNotmask + 1
				if countNotmask>25:
					flag = 1
					break
			

			color = (0, 255, 0) if label == "Mask" else (0, 0, 255)

			# include the probability in the label
			label = "{}: {:.2f}%".format(label, max(mask, withoutMask) * 100)

			# display the label and bounding box rectangle on the output
			# frame
			cv2.putText(frame, label, (startX, startY - 10),
				cv2.FONT_HERSHEY_SIMPLEX, 0.45, color, 2)
			cv2.rectangle(frame, (startX, startY), (endX, endY), color, 2)

		# show the output frame
		cv2.imshow("Frame", frame)
		key = cv2.waitKey(1) & 0xFF

		# if the `q` key was pressed, break from the loop
		if key == ord("q") or flag == 1:
			break

	# do a bit of cleanup
	cv2.destroyAllWindows()
	vs.stop()
	return flagm

def give_strID(numid):
	s = str(numid)
	r = "BT" + s[2:4]
	if s[4] == '2': r += "CSE"
	else: r += "ECE"
	r += s[7::]
	return r

def takeAttendance():
	flagA = ""
	flag = 0
	countA = 0
	countNA = 0

	recognizer = cv2.face.LBPHFaceRecognizer_create()
	recognizer.read('trainer/trainer.yml')
	faceCascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml");
	

	font = cv2.FONT_HERSHEY_SIMPLEX 

	# Initialize and start realtime video capture
	cam = cv2.VideoCapture(0)
	cam.set(3, 640) # set video widht
	cam.set(4, 480) # set video height

	# Define min window size to be recognized as a face
	minW = 0.1*cam.get(3)
	minH = 0.1*cam.get(4)

	while True:
		ret, img = cam.read()
		if ret:
			gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
			faces = faceCascade.detectMultiScale(gray, 1.5, 5, minSize = (int(minW), int(minH)))
			for(x, y, w, h) in faces:
				cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
				numid, confidence = recognizer.predict(gray[y:y+h, x:x+w])
				if (confidence < 100):
					countA = countA + 1
					
					if (countA>8):
						id = give_strID(numid)
						flag = 1
						flagA = id
						break

					else:
						id = "Null"
						countNA = countNA + 1
						if (countNA>15):
							flag=1
							flagA=id
							break
				confidence = "  {0}%".format(round(100 - confidence))
				cv2.putText(img, id, (x + 5, y - 5), font, 1, (255, 255, 255), 2)
				cv2.putText(img, str(confidence), (x + 5, y + h - 5), font, 1, (255, 255, 0), 1)
				#print("\n>> Face recognized as: {0}".format(id))
		cv2.imshow('camera_view',img) 
		k = cv2.waitKey(10) & 0xff # Press 'ESC' for exiting video
		if k == 27 or flag == 1: break

	cam.release()
	cv2.destroyAllWindows()
	return flagA

def toString(detailsList) -> str:

	detailsString = ' '.join([str(elem)+"\n" for elem in detailsList]) 
	detailsString = detailsString + "\n"
	return detailsString 

def construct_gui():

	window = tk.Tk()
	window.title("Auto-Attendance")
	window.geometry("600x300")
	first_page = tk.Frame(window, padx=20, pady=20, bg="#FE9292")
	second_page = tk.Frame(window, bg="#FE9292")
	third_page = tk.Frame(window, bg="#FE9292")


	def clear_frame(frame):
		for widget in frame.winfo_children():
			widget.destroy()
		frame.pack_forget()

	def show_info_box(list2, subjects):
		ret = mb.askyesno("Confirm Your Identity", toString(list2))
		if ret:
			global detailsList2, identityConfirmed
			detailsList2.clear()
			detailsList2.append('\nConfirmed Identity:\n\n')
			detailsList2.extend(list2)
			identityConfirmed = True
			clear_frame(third_page)
			construct_third_page()
		else:
			list2 = start1(subjects)
			show_info_box(list2, subjects)
		
	def construct_first_page():
		first_page.pack(fill="both", expand=True)

		labelFont = tkFont.Font(family="Ubuntu Mono", size=20, weight="bold", slant="italic")
		heading_label = tk.Label(first_page, text=s, font=labelFont, pady=5, padx=5)
		heading_label.pack()

		tk.Label(first_page, text="", bg="#FE9292", height=2).pack()
  
		start_button = tk.Button(first_page, text="START", relief=tk.RIDGE, bg="#243444", fg="#C7C4C2", command=construct_second_page, width=20, padx=5, pady=5, borderwidth=10)
		start_button.pack()

		exit_button = tk.Button(first_page, text="EXIT", relief=tk.RIDGE, bg="#243444", fg="#C7C4C2", command=Exit, width=20, padx=5, pady=5, borderwidth=10)
		exit_button.pack()

	def construct_second_page():

		def submit():
			
			subjects=v.get()
			subjects = sheetlist[subjects]
			clear_frame(third_page)

			list2 = start1(subjects)
			construct_third_page()
			show_info_box(list2, subjects)

		clear_frame(first_page)
		second_page.pack(fill="both", expand=True)

		labelFont = tkFont.Font(family="Ubuntu Mono", size=20, weight="bold", slant="italic")

		heading_label = tk.Label(second_page, text="Enter Subject", font=labelFont, pady=5, padx=5)
		heading_label.pack()
		
		subjects = sheetlist
		v = tk.IntVar()


		for val, subject in enumerate(subjects):
			tk.Radiobutton(second_page, text=subject, padx=260, variable=v,value=val, bg="#FE9292").pack(anchor=tk.W)


		submit_button = tk.Button(second_page, text="Submit", relief=tk.RIDGE, bg="#243444", fg="#C7C4C2", command=submit, padx=5, pady=5, borderwidth=10)
		submit_button.pack()   

	def construct_third_page():
		clear_frame(second_page)
		third_page.pack(fill="both", expand=True)

		global detailsList2, identityConfirmed
		labelFont = tkFont.Font(family="Ubuntu Mono", size=12, weight="normal", slant="italic")

		details = toString(detailsList2)
		detailsBox = tk.Label(third_page, text=details, font=labelFont, padx=5, pady=5,  bg="#FE9292")
		detailsBox.pack(fill="x")  

		if identityConfirmed:
			exit_button = tk.Button(third_page, text="Finish", relief=tk.RIDGE, bg="#243444", fg="#C7C4C2", command=finish, padx=5, pady=5, borderwidth=10, width=20)
			# tk.Label(text="", height=10)
			exit_button.pack()


	def finish():
		global identityConfirmed
		identityConfirmed = False
		clear_frame(third_page)
		construct_first_page()

	def Exit():
		window.quit()
		sys.exit(0)
	 
	construct_first_page()    
	window.mainloop()



def start1(subject): 

	while True:
		maskbit = checkMask()
		if (maskbit):
			subcode = subject
									
			counter = 0
			id = 'Null'

			while (counter<3):
				id = takeAttendance()
				if (id=='Null'):
					counter = counter + 1
				else:
					break

			if (counter == 3): # when not recognized.
				msg = 'Please enter your roll number:'
				id = input('Enter id')

			name = Trigger(id,subcode)          
			print('Hello' + id + ' : '+ name)

			list1 = [id,name]
			return(list1)
			
			
		else:
			msg = 'Please wear a mask and try again !'
			print(msg)

construct_gui()
